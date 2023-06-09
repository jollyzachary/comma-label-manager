import os
import csv
import sys
import openpyxl
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QLabel, QLineEdit, QPushButton, QFileDialog, QRadioButton, QButtonGroup, QSpinBox, QHBoxLayout, QComboBox
from PyQt5.QtWidgets import QSplashScreen
from PyQt5.QtCore import QTimer
from PyQt5.QtWidgets import QDesktopWidget
from PyQt5.QtCore import QRectF
from PyQt5.QtGui import QPainter, QFont
from PyQt5.QtCore import QEventLoop


from PyQt5.QtCore import QEventLoop

from PyQt5.QtGui import QFont

def show_splash_screen(app):
    # Load the PNG image
    pixmap = QPixmap(r'C:\\Users\\zachj\\OneDrive\\Desktop\\Comma\\assets\\comma.png')

    # Scale the image while maintaining its aspect ratio and quality
    pixmap = pixmap.scaledToWidth(700, Qt.SmoothTransformation)

    # Create a QSplashScreen with the loaded image
    splash = QSplashScreen(pixmap)

    # Set a custom font for the copyright message
    font = QFont()
    font.setBold(True)
    font.setPointSize(15)

    # Add the copyright message with the custom font
    splash.setFont(font)
    splash.showMessage("© 2023 Comma", Qt.AlignBottom | Qt.AlignHCenter, Qt.black)

    # Center the splash screen on the user's screen
    screen_geometry = app.desktop().screenGeometry()
    splash_geometry = splash.geometry()
    splash.move((screen_geometry.width() - splash_geometry.width()) // 2,
                (screen_geometry.height() - splash_geometry.height()) // 2)

    # Show the splash screen
    splash.show()

    # Create an event loop and a timer
    loop = QEventLoop()
    timer = QTimer()
    timer.timeout.connect(loop.quit)
    timer.setSingleShot(True)
    timer.start(4000)  # Close the splash screen after 3000 milliseconds (3 seconds)

    # Run the event loop
    loop.exec_()

    # Close the splash screen
    splash.close()

    return pixmap.size()
            
class GroupAndSampleManager(QWidget):
    def __init__(self, pixmap_size):
        super().__init__()

        # Set up the user interface
        self.init_ui(pixmap_size)


        # Set the window to be more translucent
        self.setWindowOpacity(0.98)

    def init_ui(self, pixmap_size):
        # Create layout and input fields
        layout = QVBoxLayout()

        font = QFont("Segoe UI")  # Set font for program

        logo = QLabel(self)
        pixmap = QPixmap('C:\\path_to_your_company_logo_here.png')
        scaled_pixmap = pixmap.scaled(85, 60) # scale the pixmap to a small size
        logo.setPixmap(scaled_pixmap)
        logo.setStyleSheet("background-color: transparent; border: 4px solid rgba(0, 0, 0, 25); border-radius: 4px;")
        layout.addWidget(logo, 0, Qt.AlignTop | Qt.AlignRight)


       # Create the radio buttons and add them to the layout
        self.single_group_rb = QRadioButton("Single group")
        self.single_group_rb.setFont(font)
        self.single_group_rb.setChecked(True)
        self.single_group_rb.toggled.connect(self.update_group_input_label)
        layout.addWidget(self.single_group_rb, 0, Qt.AlignTop | Qt.AlignLeft)

        self.multiple_groups_rb = QRadioButton("Multiple groups")
        self.multiple_groups_rb.setFont(font)
        self.multiple_groups_rb.toggled.connect(self.update_group_input_label)
        layout.addWidget(self.multiple_groups_rb, 0, Qt.AlignTop | Qt.AlignLeft)

        # Group input label and spin box
        group_input_layout = QHBoxLayout()
        

        self.group_input_label = QLabel("Enter group number:")
        group_input_layout.addWidget(self.group_input_label)
        self.group_input_label.setFont(font)

        self.group_spin_box = QSpinBox()
        self.group_spin_box.setMinimum(1)
        self.group_spin_box.setMaximum(150)
        group_input_layout.addWidget(self.group_spin_box)

        layout.addLayout(group_input_layout)

        self.samples_entry = QLineEdit()
        self.samples_entry.setPlaceholderText("Enter number of samples")
        self.samples_entry.setFont(font)
        layout.addWidget(self.samples_entry)

        self.procedures_entry = QLineEdit()
        self.procedures_entry.setPlaceholderText("Enter number of procedures")
        self.procedures_entry.setFont(font)
        layout.addWidget(self.procedures_entry)

        run_button = QPushButton("Run")
        run_button.clicked.connect(self.run_program)
        run_button.setFont(font)
        layout.addWidget(run_button)

        # Blade length selection
        blade_length_layout = QHBoxLayout()

        self.blade_length_label = QLabel("Select blade length:")
        self.blade_length_label.setStyleSheet("background-color: transparent;")
        self.blade_length_label.setFont(font)
        blade_length_layout.addWidget(self.blade_length_label)

        self.blade_length_combo = QComboBox()
        self.blade_length_combo.addItem("6\"", userData=38)
        self.blade_length_combo.addItem("9\"", userData=58)
        self.blade_length_combo.addItem("12\"", userData=70)
        self.blade_length_combo.addItem("Custom", userData=None)
        self.blade_length_combo.setFont(font)
        blade_length_layout.addWidget(self.blade_length_combo)

        self.blade_length_combo.setStyleSheet("""
            QComboBox {
                background-color: white;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                selection-background-color: grey;
            }
        """)


        # Custom spaces input
        self.custom_spaces_label = QLabel("Custom spaces:")
        self.custom_spaces_label.setStyleSheet("background-color: transparent;")
        self.custom_spaces_label.setDisabled(True)
        self.custom_spaces_label.setFont(font)
        blade_length_layout.addWidget(self.custom_spaces_label)


        self.custom_spaces_spin_box = QSpinBox()
        self.custom_spaces_spin_box.setStyleSheet("background-color: white;")
        self.custom_spaces_spin_box.setMinimum(0)
        self.custom_spaces_spin_box.setMaximum(150)
        self.custom_spaces_spin_box.setValue(38) # Sets the deault value to 38
        self.custom_spaces_spin_box.setStyleSheet("QSpinBox {color: lightgrey; background-color: white;}")
        self.custom_spaces_spin_box.setDisabled(True)
        self.custom_spaces_spin_box.setFont(font)
        blade_length_layout.addWidget(self.custom_spaces_spin_box)

        self.blade_length_combo.currentIndexChanged.connect(self.update_custom_spaces)

       
        layout.addLayout(blade_length_layout)


        # Set the background color of the widget
        self.setStyleSheet("background-color: rgba(255, 15, 15, 200);")
        self.samples_entry.setStyleSheet("background-color: white;")
        self.procedures_entry.setStyleSheet("background-color: white;")
        self.group_spin_box.setStyleSheet("background-color: white;")
        self.single_group_rb.setStyleSheet("background-color: transparent;")
        self.multiple_groups_rb.setStyleSheet("background-color: transparent;")
        self.group_input_label.setStyleSheet("background-color: transparent;")
        run_button.setStyleSheet("background-color: white;")

        self.setLayout(layout)
        self.setWindowTitle("Comma Label Manager")
        self.setGeometry(100, 100, pixmap_size.width(), pixmap_size.height())
        self.center_on_screen()
    
        self.show()

    def center_on_screen(self):
        screen_geometry = QDesktopWidget().screenGeometry()
        window_geometry = self.frameGeometry()
        window_geometry.moveCenter(screen_geometry.center())
        self.move(window_geometry.topLeft())



    def update_custom_spaces(self, index):
        selected_text = self.blade_length_combo.itemText(index)
        if selected_text == "Custom":
            self.custom_spaces_label.setDisabled(False)
            self.custom_spaces_spin_box.setDisabled(False)
            self.custom_spaces_spin_box.setStyleSheet("QSpinBox {color: black;}")
            self.custom_spaces_spin_box.setStyleSheet("background-color: white;")
        else:
            spaces = self.blade_length_combo.itemData(index)
            self.custom_spaces_spin_box.setValue(spaces)
            self.custom_spaces_spin_box.setStyleSheet("background-color: white;")
            self.custom_spaces_label.setDisabled(True)
            self.custom_spaces_spin_box.setDisabled(True)
            self.custom_spaces_spin_box.setStyleSheet("QSpinBox {color: lightgrey; background-color: white;}")


    def update_group_input_label(self):
        if self.single_group_rb.isChecked():
            self.group_input_label.setText("Enter group number:")
        else:
            self.group_input_label.setText("Enter number of groups:")

    def run_program(self):
        if self.single_group_rb.isChecked():
            group_list = [self.group_spin_box.value()]
        else:  # multiple_groups_rb is checked
            num_groups = self.group_spin_box.value()
            group_list = list(range(1, num_groups + 1))


        num_samples = int(self.samples_entry.text())
        num_procedures = int(self.procedures_entry.text())

        if self.blade_length_combo.currentText() == "Custom":
            spaces = self.custom_spaces_spin_box.value()
        else:
                spaces = self.blade_length_combo.currentData()

        # Create a new workbook and add a worksheet named 'Output'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Output"

        # Starting row number
        row_num = 1

        # Loop through each group
        for group_num in group_list:

            # Loop through each procedure
            for p in range(1, num_procedures + 1):

                # Loop through each sample in the group
                for s in range(1, num_samples + 1):

                    # Write the row header in the first cell with spaces
                    ws.cell(row=row_num, column=1).value = f"{group_num}-{p:2}{s:02}" + " " * spaces + f"{group_num}-{p:2}{s:02}"
                                     
                    # Increment the row number
                    row_num += 1

        # Save the workbook to a file
        file_name, _ = QFileDialog.getSaveFileName(self, "Save output", "", "CSV files (*.csv)")
        if file_name:
            if not file_name.endswith(".csv"):
                file_name += ".csv"

            # Create a list of rows containing the values from the cells in the workbook
            data = []
            for row in ws.iter_rows():
                data.append([cell.value for cell in row])

            # Save the workbook data to a CSV file
            with open(file_name, 'w', newline='') as file:
                writer = csv.writer(file)
                for row in data:
                    writer.writerow(row)

            # Open the saved file using the default program
            try:
                os.startfile(file_name)
            except FileNotFoundError:
                print(f"Error: Failed to open the file {file_name}")
        else:
            print("No file selected.")

def main():
    app = QApplication(sys.argv)

    pixmap_size = show_splash_screen(app)

    group_and_sample_manager = GroupAndSampleManager(pixmap_size)
    group_and_sample_manager.show()

    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
