import streamlit as st
import csv
import openpyxl
import os

# Set up the Streamlit app
st.title('Comma Label Manager')

# Create layout and input fields
single_group_rb = st.radio('Grouping:', ('Single group', 'Multiple groups'))
group_input_label = st.text_input('Enter group number:' if single_group_rb == 'Single group' else 'Enter number of groups:')
samples_entry = st.text_input('Enter number of samples:')
procedures_entry = st.text_input('Enter number of procedures:')
blade_length_combo = st.selectbox('Select blade length:', ('6"', '9"', '12"', 'Custom'))
custom_spaces_spin_box = st.number_input('Custom spaces:', value=38, min_value=0, max_value=150, step=1, key='custom_spaces')

# Run button
if st.button('Run'):

    # Get the selected options and input values
    single_group = single_group_rb == 'Single group'
    group_input = int(group_input_label)
    num_samples = int(samples_entry)
    num_procedures = int(procedures_entry)
    blade_length = blade_length_combo
    custom_spaces = int(custom_spaces_spin_box)

    # Create a new workbook and add a worksheet named 'Output'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Output"

    # Starting row number
    row_num = 1

    # Loop through each group
    for group_num in range(1, group_input + 1) if not single_group else [group_input]:

        # Loop through each procedure
        for p in range(1, num_procedures + 1):

            # Loop through each sample in the group
            for s in range(1, num_samples + 1):

                # Write the row header in the first cell with spaces
                ws.cell(row=row_num, column=1).value = f"{group_num}-{p:2}{s:02}" + " " * custom_spaces + f"{group_num}-{p:2}{s:02}"

                # Increment the row number
                row_num += 1

    # Save the workbook to a file
    file_name = st.text_input('Enter file name:', key='output_file_name')
    if st.button('Save output') and file_name:
        if not file_name.endswith(".csv"):
            file_name += ".csv"

        # Create a list of rows containing the values from the cells in the workbook
        data = []
        for row in ws.iter_rows():
            data.append([cell.value for cell in row])

        # Save the workbook data to a CSV file
        with open(file_name, 'w', newline='') as file:
            writer = csv.writer(file)
            for row in data:
                writer.writerow(row)

        # Open the saved file using the default program
        try:
            os.startfile(file_name)
        except FileNotFoundError:
            st.error(f"Error: Failed to open the file {file_name}")
